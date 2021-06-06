# 난는 교수입니다
# 수업을 하는데 점수 비중이 이렇읍니다..

# -출석       10
# -퀴즈1      10
# -퀴즈2      10
# -중간       20
# -기말       30
# -프로젝트    20
# -총합       100

# 마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서
# 퀴즈2 문제에 오류를 발견하여 모두 만점처리를 하기로 했습니다.
# 현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오.

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(sum), I열에 성적 정보 추가
#   총점  >= 90 == A
#        >= 80 == B
#        >= 70 == C
#        <  70 == D
# 3. 출석이 5 미만인 경우, 총점 불문하고 F

# 최종 파일명: scores.xlsx

# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간, 기말, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
wb = Workbook()
ws = wb.active

#데이터 입력
ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
# ws.append([1,10,8,5,14,26,12])
# ws.append([2,7,3,7,15,24,18])
# ws.append([3,9,5,8,8,12,4])
# ws.append([4,7,8,7,17,21,18])
# ws.append([5,7,8,7,16,25,15])
# ws.append([6,3,5,8,8,17,0])
# ws.append([7,4,9,10,16,27,18])
# ws.append([8,6,6,6,15,19,17])
# ws.append([9,10,10,9,19,30,19])
# ws.append([10,9,8,8,20,25,20])
# 좀 더 간략하게 작성하자면...
scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]
for s in scores:
    ws.append(s)

#row 및 column 크기 조정
ws.row_dimensions[1].height = 50
ws.column_dimensions["A"].width = 5

# 1. 퀴즈2 점수를 10으로 수정
# for num in range(2, ws.max_row+1):
#     ws.cell(row=num, column=4, value=10)

# another 1)
for idx, cell in enumerate(ws["D"]):
    if idx == 0:        # 제목인 경우 skip, idx=index(문자)
        continue
    cell.value = 10

# 2. H열에 총점(sum), I열에 성적 정보 추가
# ws.cell(row=1, column=ws.max_column+1, value="총점")  #ws["H1"] = "총점"
# ws.cell(row=1, column=ws.max_column+1, value="성적")  #ws["I1"] = "성적"
#
# for i in range(2, ws.max_row+1):
#     ws.cell(row=i, column=ws.max_column-1, value="=SUM(B{0}:G{0})".format(i))
#     # 3. 출석이 5 미만인 경우, 총점 불문하고 F
#     ws.cell(row=i, column=ws.max_column, value=r'=IF(B{0}<5,"F", IF(H{0}>=90,"A", IF(H{0}>=80, "B", IF(H{0}>=70, "C","D"))))'.format(i))
# for row in ws.iter_rows(min_col=2, max_col=2):
#     for cell in row:
#         if isinstance(cell.value, int) and cell.value < 5:
#             cell.fill = PatternFill(fgColor="FF0066", fill_type="solid")
#             cell.font = Font(color="ffffff")

# another 2)
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):   #2열부터 시작
    sum_val = sum(score[1:]) - score[3] + 10    #총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)      #H열(8)

    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    #another 3)
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade

#테두리 정의
left_border = Border(left=Side(style="thin"))
right_border = Border(right=Side(style="thin"))
bottom_border = Border(bottom=Side(style="double"))

#첫 row 가운데 정렬 및 테두리 적용
for cell in ws["A"]:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # cell.border = right_border
for cell in ws["I"]:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # cell.border = left_border
for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.border = bottom_border
# ws["A1"].alignment = Alignment(horizontal="center", vertical="bottom")
# ws["I1"].alignment = Alignment(horizontal="center", vertical="bottom")
for cell in ws[ws.max_row]:
    cell.border = bottom_border

#팬 고정
ws.freeze_panes = "B2"

wb.save("scores.xlsx")