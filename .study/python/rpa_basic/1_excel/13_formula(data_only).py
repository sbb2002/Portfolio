from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# 코딩 당시 사용한 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

# 애초에 data_only를 쓰면 data를 그대로 가져오는데...
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)  # None이 나옴. 사실은...
        # ws["A2"] = "=SUM(1, 2, 3)"
        # ws["A3"] = "=AVERAGE(1, 2, 3)"
        # 로 썼기에 얘네가 실행되기 전 상태로 로드하기 때문임.
        # evalueate 되지 않으면 None이라고 함.

        # 해결법은 formula.xlsx를 켜서 바로 끄면 저장하겠냐고 묻는데
        # 저장하면 함수결괏값이 정수화됨을 노림.
        # 이 과정을 한 후, 이 코딩을 돌리면 None없이 출력됨.
