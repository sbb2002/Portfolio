from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 이라고 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])         #참조
print(ws["A1"].value)   #값
print(ws["A10"].value)   #값 없음

# row = 1, 2, 3, 4, ...
# column = A(1), B(2), C(3), D(4), ...
print(ws.cell(row=1, column=1).value)       #A1
print(ws.cell(row=1, column=2).value)       #B1

c = ws.cell(column=3, row=1, value=10)          #ws["C1"] = 10 과 동일
print(c.value)      # ws["C1"].value 와 동일

from random import *

#반복문을 이용해서 랜덤 숫자 채우기
# index = 1
# for x in range(1, 11):      #10개 rows
#     for y in range(1, 11):  #10개 columns
#         # ws.cell(row=x, column=y, value=randint(0, 100))     # 0 ~ 100
#         ws.cell(row=x, column=y, value=index)
#         index += 1
wb.save("sample_cell.xlsx")