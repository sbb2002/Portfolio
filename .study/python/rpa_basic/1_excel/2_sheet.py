from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()      #새로운 sheet 생성 (기본이름)
ws.title = "MySheet"
ws.sheet_properties.tabColor = "9999ff"   #RGB 참조: https://www.w3schools.com/colors/colors_rgb.asp

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)    #2번쨰 index에 sheet 생성

new_ws = wb["NewSheet"]       #dict로 sheet 여러개 생성 가능

print(wb.sheetnames)        #모든 sheet이름 확인

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"
wb.save("sample.xlsx")