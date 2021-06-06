from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호, 영어, 수학
a1 = ws["A1"]   #번호
b1 = ws["B1"]   #영어
c1 = ws["C1"]   #수학

ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50

a1.font = Font(color="FF0000", italic=True, bold=True)  #red, italy, bold
b1.font = Font(color="CC33FF", name="Arial", strike=True)   #Arial, 취소선
c1.font = Font(color="0000FF", size=20, underline="single") #size 20, 밑줄

#테두리
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"), top=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

#>=90 인 학생을 초록으로 강조
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")  #정렬
        if cell.column == 1:    #A열 제외
            continue

        #셀이 정수형이면서 90점 이상일 때
        if isinstance(cell.value, int) and cell.value >= 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

#틀 고정
ws.freeze_panes = "B2"  #B2 기준으로 고정되서 스크롤을 상하좌우로 이동해도 각 첫 row, col은 고정됨. 굳


wb.save("sample_style.xlsx")