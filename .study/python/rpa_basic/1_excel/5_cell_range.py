from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])   #column = A, B, C
for i in range(1, 11):  #10줄
    ws.append([i, randint(1, 100), randint(1, 100)])

# #영어 센세가 영어성적만 가지고 오려함
# col_B = ws["B"]
# # print(col_B)
# for cell in col_B:
#     print(cell.value)
#
# #담임 센세가 영어, 수학 성적을 가지고 오려함
# col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)
#
# row_title = ws[1]   #2nd row
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]     #2~6줄 (1~5번학생 data)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string
#
# print("모든 학생의 번호, 영어, 수학 성적입니다.")
# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") #셀 좌표
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()


# # 전체 rows
# # print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     # print(row[0].value)     #학생번호
#     print(row[1].value)     #반 전체 영어성적
#     # print(row[2].value)     #반 전체 수학성적
#
# # 전체 columns
# for column in tuple(ws.columns):
#     print(column[0].value)
# # print(tuple(ws.columns))

# for row in ws.iter_rows():  #전체 row
#     print(row[1].value)

# for column in ws.iter_cols(): #전체 column
#     print(column[0].value)

# for row in ws.iter_rows(min_row=1, max_row=5):  #엑셀 기준 1~5줄 = 1~4번의 영어성적
#     print(row[1].value)

#1~11번째 줄, 2~3번째 열
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):   #한 학생의 성적
    print(row[0].value, row[1].value)     #수학, 영어
    # print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):    #한 과목의 성적
    print(col)


wb.save("sample_cell_range.xlsx")